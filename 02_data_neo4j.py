import os
from app import app as apps,db,app_tools
from app.user import models
from app.tools.neo4j_handler import graph_handler

from py2neo import Graph, Node, Relationship

from openpyxl import load_workbook

db.create_all()

class NodeObj:
    def __init__(self, label="", name="", property={}):
        self.node = Node(label, name=name, **property)
        self.label = label
        self.name = name
        self.property = property

class RelObj:
    def __init__(self, node1: NodeObj, name, node2: NodeObj, property={}):
        self.node1 = node1
        self.node2 = node2
        self.name = name
        self.property = property

graph_handler.clear_db()

def write_db():
    path = f"生物交叉领域术语表.xlsx"
    wb = load_workbook(path)

    # 选择工作表'Sheet1'
    sheet = wb['Sheet1']

    # 使用iter_rows()方法逐行读取数据
    level_01 = ["", ""]
    level_02 = ["", ""]
    level_03 = ["", ""]
    level_04 = ["", ""]
    level_05 = ["", ""]

    node_lis = []
    rel_lis = []

    for index, row in enumerate(sheet.iter_rows(values_only=True)):
        if index == 0:
            continue
        for i, cell in enumerate(row):
            l, r = divmod(i, 2)

            if cell is None:
                cell = ""

            if l == 3:
                level_04[r] = cell
            elif l == 4:
                level_05[r] = cell

            if cell == "":
                continue

            if l == 0:
                level_01[r] = cell
            elif l == 1:
                level_02[r] = cell
            elif l == 2:
                level_03[r] = cell

        if level_01[0]:
            Node01 = NodeObj("一级", name=level_01[0], property={"en": level_01[1]})
            node_lis.append(Node01)
        if level_02[0]:
            Node02 = NodeObj("二级", name=level_02[0], property={"en": level_02[1]})
            node_lis.append(Node02)
            rel_lis.append(RelObj(Node01, "下级", Node02))
        if level_03[0]:
            Node03 = NodeObj("三级", name=level_03[0], property={"en": level_03[1]})
            node_lis.append(Node03)
            rel_lis.append(RelObj(Node02, "下级", Node03))
        if level_04[0]:
            Node04 = NodeObj("四级", name=level_04[0], property={"en": level_04[1]})
            node_lis.append(Node04)
            rel_lis.append(RelObj(Node03, "下级", Node04))
        if level_05[0]:
            Node05 = NodeObj("五级", name=level_05[0], property={"en": level_05[1]})
            node_lis.append(Node05)
            rel_lis.append(RelObj(Node04, "下级", Node05))

    # 关闭工作簿
    wb.close()

    tx = graph_handler.graph.begin()
    for node in node_lis:
        tx.merge(node.node, node.label, "name")
    tx.commit()

    tx = graph_handler.graph.begin()
    for rel in rel_lis:
        tx.create(Relationship(rel.node1.node, rel.name, rel.node2.node, **rel.property))
    tx.commit()

if __name__ == "__main__":
    write_db()